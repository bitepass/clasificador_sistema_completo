from flask import Blueprint, request, jsonify, current_app, send_file
from werkzeug.utils import secure_filename
import pandas as pd
import os
import json
import time
import threading
from datetime import datetime, date
import traceback
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import google.generativeai as genai
import openai
import re

from src.models.models import db, Planilla, HechoDelictivo, LogProcesamiento

clasificador_bp = Blueprint('clasificador', __name__)

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Variables globales para control de proceso
procesos_activos = {}

# Configurar APIs
def configurar_apis():
    """Configura las APIs de Gemini y OpenAI"""
    try:
        gemini_key = os.getenv('GEMINI_API_KEY')
        if gemini_key and gemini_key != 'your_gemini_api_key_here':
            genai.configure(api_key=gemini_key)
        
        openai_key = os.getenv('OPENAI_API_KEY')
        if openai_key and openai_key != 'your_openai_api_key_here':
            openai.api_key = openai_key
    except Exception as e:
        logger.error(f"Error configurando APIs: {str(e)}")

# Categorías de clasificación
CATEGORIAS_CLASIFICACION = {
    'JURISDICCION': ['URBANA', 'RURAL', 'MIXTA', 'OTROS'],
    'CALIFICACION': [
        'HOMICIDIO_SIMPLE', 'FEMICIDIO', 'INTRAFAMILIAR', 'EN_RINA', 'EN_OCASION_DE_ROBO', 'AJUSTE_DE_CUENTAS',
        'ROBO_ESCRUCHE', 'ROBO_ENTRADERA', 'ROBO_ASALTO_FINCA', 'ROBO_ASALTO_VIA_PUBLICA', 'ROBO_ASALTO_COMERCIO',
        'ROBO_MOTOCHORROS', 'ROBO_ROBACABLES', 'ROBO_SIMPLE', 'ROBO_ROBARRUEDAS', 'ROBO_ROMPEVIDRIOS',
        'ROBO_ARREBATADOR', 'ROBO_BICICLETA', 'ROBO_CHOFERES_REPARTIDORES',
        'ENFRENTAMIENTOS_EN_OCASION_DE_ROBO', 'ENFRENTAMIENTOS_AJUSTE_DE_CUENTAS', 'ENFRENTAMIENTOS_PROCEDIMIENTO_POLICIAL',
        'ENFRENTAMIENTOS_EN_RINA', 'ENFRENTAMIENTOS_BANDAS_ANTAGONICAS',
        'SUSTRACCION_AUTOMOTOR_LEVANTAMIENTO', 'SUSTRACCION_AUTOMOTOR_ASALTO', 'SUSTRACCION_MOTOVEHICULO_LEVANTAMIENTO',
        'SUSTRACCION_MOTOVEHICULO_ASALTO', 'LESIONES_ARMA_FUEGO', 'LESIONES_ARMA_BLANCA', 'LESIONES_ARMA_IMPROPIA',
        'USURPACION', 'ABUSO_SEXUAL_SIMPLE', 'ABUSO_SEXUAL_ACCESO_CARNAL', 'LEY_23737_TENENCIA', 'LEY_23737_CONSUMO',
        'LEY_23737_COMERCIALIZACION', 'LEY_23737_SIEMBRA', 'ABIGEATO', 'ESTAFAS_MARKETPLACE', 'ESTAFAS_WHATSAPP',
        'ESTAFAS_CUENTO_DEL_TIO', 'ESTAFAS_OTROS', 'ABUSO_DE_ARMAS', 'TENENCIA_DE_ARMAS', 'PORTACION_DE_ARMAS',
        'ENCUBRIMIENTO_VIA_PUBLICA', 'ENCUBRIMIENTO_TALLER', 'ENCUBRIMIENTO_DOMICILIO_PARTICULAR', 'NINGUNO_DE_INTERES'
    ],
    'MODALIDAD': [
        'MODALIDAD_HOMICIDIO_SIMPLE', 'MODALIDAD_FEMICIDIO', 'MODALIDAD_INTRAFAMILIAR', 'MODALIDAD_EN_RINA',
        'MODALIDAD_EN_OCASION_DE_ROBO', 'MODALIDAD_AJUSTE_DE_CUENTAS', 'MODALIDAD_ROBO_ESCRUCHE',
        'MODALIDAD_ROBO_ENTRADERA', 'MODALIDAD_ROBO_ASALTO_FINCA', 'MODALIDAD_ROBO_ASALTO_VIA_PUBLICA',
        'MODALIDAD_ROBO_ASALTO_COMERCIO', 'MODALIDAD_ROBO_MOTOCHORROS', 'MODALIDAD_ROBO_ROBACABLES',
        'MODALIDAD_ROBO_SIMPLE', 'MODALIDAD_ROBO_ROBARRUEDAS', 'MODALIDAD_ROBO_ROMPEVIDRIOS',
        'MODALIDAD_ROBO_ARREBATADOR', 'MODALIDAD_ROBO_BICICLETA', 'MODALIDAD_ROBO_CHOFERES_REPARTIDORES',
        'MODALIDAD_ENFRENTAMIENTOS_EN_OCASION_DE_ROBO', 'MODALIDAD_ENFRENTAMIENTOS_AJUSTE_DE_CUENTAS',
        'MODALIDAD_ENFRENTAMIENTOS_PROCEDIMIENTO_POLICIAL', 'MODALIDAD_ENFRENTAMIENTOS_EN_RINA',
        'MODALIDAD_ENFRENTAMIENTOS_BANDAS_ANTAGONICAS', 'MODALIDAD_SUSTRACCION_AUTOMOTOR_LEVANTAMIENTO',
        'MODALIDAD_SUSTRACCION_AUTOMOTOR_ASALTO', 'MODALIDAD_SUSTRACCION_MOTOVEHICULO_LEVANTAMIENTO',
        'MODALIDAD_SUSTRACCION_MOTOVEHICULO_ASALTO', 'MODALIDAD_LESIONES_ARMA_FUEGO',
        'MODALIDAD_LESIONES_ARMA_BLANCA', 'MODALIDAD_LESIONES_ARMA_IMPROPIA', 'OTROS'
    ],
    'VICTIMAS': ['FEMENINO', 'MASCULINO', 'AMBOS', 'OTROS'],
    'LESIONADO': ['SI', 'NO', 'OTROS'],
    'IMPUTADOS': ['FEMENINO', 'MASCULINO', 'AMBOS', 'OTROS'],
    'EDAD': ['MAYOR', 'MENOR', 'AMBOS', 'OTROS'],
    'ARMAS': ['FUEGO', 'BLANCA', 'IMPROPIA', 'NINGUNA', 'OTROS'],
    'LUGAR': ['FINCA', 'VIA_PUBLICA', 'COMERCIO', 'ESTABLECIMIENTO_EDUCATIVO', 'OTROS'],
    'TENTATIVA': ['SI', 'NO', 'OTROS']
}

def clasificar_texto_con_gemini(texto):
    """Clasificar texto usando Gemini API"""
    try:
        model = genai.GenerativeModel('gemini-pro')
        
        prompt = f"""
        Analiza el siguiente texto de un hecho delictivo y clasifícalo según estas categorías exactas:

        JURISDICCION: {', '.join(CATEGORIAS_CLASIFICACION['JURISDICCION'])}
        CALIFICACION: {', '.join(CATEGORIAS_CLASIFICACION['CALIFICACION'])}
        MODALIDAD: {', '.join(CATEGORIAS_CLASIFICACION['MODALIDAD'])}
        VICTIMAS: {', '.join(CATEGORIAS_CLASIFICACION['VICTIMAS'])}
        LESIONADO: {', '.join(CATEGORIAS_CLASIFICACION['LESIONADO'])}
        IMPUTADOS: {', '.join(CATEGORIAS_CLASIFICACION['IMPUTADOS'])}
        EDAD: {', '.join(CATEGORIAS_CLASIFICACION['EDAD'])}
        ARMAS: {', '.join(CATEGORIAS_CLASIFICACION['ARMAS'])}
        LUGAR: {', '.join(CATEGORIAS_CLASIFICACION['LUGAR'])}
        TENTATIVA: {', '.join(CATEGORIAS_CLASIFICACION['TENTATIVA'])}

        Texto a analizar: {texto}

        Responde SOLO en formato JSON válido con estas claves exactas:
        {{
            "jurisdiccion": "valor",
            "calificacion": "valor",
            "modalidad": "valor",
            "victimas": "valor",
            "lesionado": "valor",
            "imputados": "valor",
            "edad": "valor",
            "armas": "valor",
            "lugar": "valor",
            "tentativa": "valor",
            "observaciones": "texto libre con detalles relevantes"
        }}
        """
        
        response = model.generate_content(prompt)
        resultado = json.loads(response.text.strip())
        return resultado, 'GEMINI'
        
    except Exception as e:
        logger.error(f"Error en clasificación con Gemini: {str(e)}")
        return None, 'ERROR_GEMINI'

def clasificacion_basada_en_reglas(texto):
    """Clasificación básica basada en reglas como alternativa"""
    try:
        texto_lower = texto.lower()
        
        # Reglas básicas de clasificación (usar las claves correctas en mayúsculas)
        clasificacion = {
            'JURISDICCION': 'URBANA',  # Por defecto urbana
            'CALIFICACION': 'OTROS',
            'MODALIDAD': 'OTROS',
            'VICTIMAS': 'OTROS',
            'LESIONADO': 'OTROS',
            'IMPUTADOS': 'OTROS',
            'EDAD': 'OTROS',
            'ARMAS': 'NINGUNA',
            'LUGAR': 'OTROS',
            'TENTATIVA': 'NO',
            'observaciones': 'Clasificación automática por reglas'
        }
        
        # Reglas para calificación (más completas)
        if any(word in texto_lower for word in ['homicidio', 'asesinato', 'muerte', 'mató', 'muerto', 'cadáver']):
            clasificacion['CALIFICACION'] = 'HOMICIDIO_SIMPLE'
        elif any(word in texto_lower for word in ['femicidio', 'violencia de género']):
            clasificacion['CALIFICACION'] = 'FEMICIDIO'
        elif any(word in texto_lower for word in ['robo', 'asalto', 'sustracción', 'hurto', 'robó', 'sustrajo']):
            clasificacion['CALIFICACION'] = 'ROBO_SIMPLE'
            if any(word in texto_lower for word in ['vía pública', 'calle', 'vereda']):
                clasificacion['CALIFICACION'] = 'ROBO_ASALTO_VIA_PUBLICA'
            elif any(word in texto_lower for word in ['comercio', 'negocio', 'tienda']):
                clasificacion['CALIFICACION'] = 'ROBO_ASALTO_COMERCIO'
        elif any(word in texto_lower for word in ['lesiones', 'herido', 'golpeó', 'agredió']):
            clasificacion['CALIFICACION'] = 'LESIONES_ARMA_IMPROPIA'
        elif any(word in texto_lower for word in ['estafa', 'defraudación', 'engaño']):
            clasificacion['CALIFICACION'] = 'ESTAFAS_OTROS'
        elif any(word in texto_lower for word in ['droga', 'marihuana', 'cocaína', 'estupefaciente']):
            clasificacion['CALIFICACION'] = 'LEY_23737_TENENCIA'
        
        # Reglas para armas
        if any(word in texto_lower for word in ['arma de fuego', 'pistola', 'revolver', 'bala', 'disparo', 'tiro']):
            clasificacion['ARMAS'] = 'FUEGO'
        elif any(word in texto_lower for word in ['cuchillo', 'navaja', 'arma blanca', 'puñal']):
            clasificacion['ARMAS'] = 'BLANCA'
        elif any(word in texto_lower for word in ['palo', 'bate', 'piedra', 'botella']):
            clasificacion['ARMAS'] = 'IMPROPIA'
        
        # Reglas para lugar
        if any(word in texto_lower for word in ['calle', 'vía pública', 'vereda', 'avenida', 'ruta', 'plaza']):
            clasificacion['LUGAR'] = 'VIA_PUBLICA'
        elif any(word in texto_lower for word in ['casa', 'domicilio', 'vivienda', 'hogar', 'departamento']):
            clasificacion['LUGAR'] = 'FINCA'
        elif any(word in texto_lower for word in ['comercio', 'negocio', 'tienda', 'local', 'shop']):
            clasificacion['LUGAR'] = 'COMERCIO'
        elif any(word in texto_lower for word in ['escuela', 'colegio', 'universidad', 'instituto']):
            clasificacion['LUGAR'] = 'ESTABLECIMIENTO_EDUCATIVO'
        
        # Reglas para víctimas
        if any(word in texto_lower for word in ['mujer', 'femenina', 'señora', 'chica']):
            clasificacion['VICTIMAS'] = 'FEMENINO'
        elif any(word in texto_lower for word in ['hombre', 'masculino', 'señor', 'chico']):
            clasificacion['VICTIMAS'] = 'MASCULINO'
        
        # Reglas para lesionado
        if any(word in texto_lower for word in ['herido', 'lesionado', 'golpeado', 'lastimado']):
            clasificacion['LESIONADO'] = 'SI'
        
        return clasificacion, 'REGLAS'
        
    except Exception as e:
        logger.error(f"Error en clasificación por reglas: {str(e)}")
        return None, 'ERROR_REGLAS'

def clasificacion_por_defecto():
    """Clasificación por defecto cuando todo falla"""
    return {
        'jurisdiccion': 'OTROS',
        'calificacion': 'NINGUNO_DE_INTERES',
        'modalidad': 'OTROS',
        'victimas': 'OTROS',
        'lesionado': 'OTROS',
        'imputados': 'OTROS',
        'edad': 'OTROS',
        'armas': 'NINGUNA',
        'lugar': 'OTROS',
        'tentativa': 'NO',
        'observaciones': 'Sin información suficiente para clasificar'
    }, 'DEFECTO'

def limpiar_clasificacion(clasificacion):
    """Valida y limpia la clasificación obtenida"""
    if not isinstance(clasificacion, dict):
        return None
    
    clasificacion_limpia = {}
    
    for categoria, valor in clasificacion.items():
        if categoria in CATEGORIAS_CLASIFICACION:
            if valor in CATEGORIAS_CLASIFICACION[categoria]:
                clasificacion_limpia[categoria] = valor
            else:
                clasificacion_limpia[categoria] = 'OTROS'
        elif categoria == 'observaciones':
            clasificacion_limpia[categoria] = str(valor)[:500]  # Limitar longitud
    
    # Asegurar que todas las categorías estén presentes
    for categoria in CATEGORIAS_CLASIFICACION.keys():
        if categoria not in clasificacion_limpia:
            clasificacion_limpia[categoria] = 'OTROS'
    
    if 'observaciones' not in clasificacion_limpia:
        clasificacion_limpia['observaciones'] = ''
    
    return clasificacion_limpia

def clasificar_fila_con_cascada(texto, fila_num):
    """Clasifica una fila usando el sistema de cascada"""
    inicio_tiempo = time.time()
    
    # Si no hay texto, usar clasificación por defecto
    if not texto or str(texto).strip() == '' or pd.isna(texto):
        clasificacion, metodo = clasificacion_por_defecto()
        tiempo_ms = int((time.time() - inicio_tiempo) * 1000)
        return clasificacion, metodo, tiempo_ms, 0.0
    
    # Intentar con Gemini primero
    clasificacion, metodo = clasificar_texto_con_gemini(str(texto))
    if clasificacion:
        clasificacion_limpia = limpiar_clasificacion(clasificacion)
        if clasificacion_limpia:
            tiempo_ms = int((time.time() - inicio_tiempo) * 1000)
            return clasificacion_limpia, metodo, tiempo_ms, 0.9
    
    # Si Gemini falla, usar reglas
    clasificacion, metodo = clasificacion_basada_en_reglas(str(texto))
    if clasificacion:
        clasificacion_limpia = limpiar_clasificacion(clasificacion)
        if clasificacion_limpia:
            tiempo_ms = int((time.time() - inicio_tiempo) * 1000)
            return clasificacion_limpia, metodo, tiempo_ms, 0.5
    
    # Usar clasificación por defecto
    clasificacion, metodo = clasificacion_por_defecto()
    tiempo_ms = int((time.time() - inicio_tiempo) * 1000)
    return clasificacion, metodo, tiempo_ms, 0.1

@clasificador_bp.route('/upload', methods=['POST'])
def upload_file():
    """Endpoint para subir archivo Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No se encontró archivo'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Solo se permiten archivos Excel (.xlsx, .xls)'}), 400
        
        # Guardar archivo
        filename = secure_filename(file.filename)
        upload_path = os.path.join('/tmp', filename)
        file.save(upload_path)
        
        # Leer Excel para obtener información básica
        df = pd.read_excel(upload_path)
        total_filas = len(df)
        
        # Crear registro en base de datos
        planilla = Planilla(
            nombre_archivo=filename,
            total_registros=total_filas,
            estado='SUBIDO'
        )
        db.session.add(planilla)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'planilla_id': planilla.id,
            'filename': filename,
            'total_filas': total_filas,
            'message': 'Archivo subido correctamente'
        })
        
    except Exception as e:
        logger.error(f"Error subiendo archivo: {str(e)}")
        return jsonify({'error': f'Error procesando archivo: {str(e)}'}), 500

@clasificador_bp.route('/process/<int:planilla_id>', methods=['POST'])
def process_planilla(planilla_id):
    """Endpoint para procesar archivo con clasificación"""
    try:
        configurar_apis()
        
        planilla = Planilla.query.get_or_404(planilla_id)
        
        if planilla.estado == 'PROCESANDO':
            return jsonify({'error': 'El archivo ya se está procesando'}), 400
        
        # Marcar como procesando
        planilla.estado = 'PROCESANDO'
        planilla.registros_procesados = 0
        db.session.commit()
        
        # Iniciar procesamiento en hilo separado
        thread = threading.Thread(
            target=procesar_archivo_async,
            args=(planilla_id,)
        )
        thread.daemon = True
        thread.start()
        
        # Guardar referencia del proceso
        procesos_activos[planilla_id] = {
            'thread': thread,
            'cancelado': False,
            'inicio': datetime.now()
        }
        
        return jsonify({
            'success': True,
            'message': 'Procesamiento iniciado',
            'planilla_id': planilla_id
        })
        
    except Exception as e:
        logger.error(f"Error iniciando procesamiento: {str(e)}")
        return jsonify({'error': f'Error: {str(e)}'}), 500

def procesar_archivo_async(planilla_id):
    """Procesa el archivo de forma asíncrona"""
    try:
        with current_app.app_context():
            planilla = Planilla.query.get(planilla_id)
            if not planilla:
                return
            
            # Leer archivo Excel
            upload_path = os.path.join('/tmp', planilla.nombre_archivo)
            df = pd.read_excel(upload_path)
            
            # Procesar cada fila
            for index, row in df.iterrows():
                # Verificar si fue cancelado
                if planilla_id in procesos_activos and procesos_activos[planilla_id]['cancelado']:
                    planilla.estado = 'CANCELADO'
                    db.session.commit()
                    return
                
                # Obtener texto para clasificar (columna 'relato' o similar)
                texto_relato = ''
                for col in ['relato', 'descripcion', 'detalle', 'observaciones']:
                    if col in df.columns:
                        texto_relato = str(row[col]) if pd.notna(row[col]) else ''
                        break
                
                # Clasificar fila
                clasificacion, metodo, tiempo_ms, confianza = clasificar_fila_con_cascada(texto_relato, index)
                
                # Crear registro en base de datos
                hecho = HechoDelictivo(
                    planilla_id=planilla_id,
                    id_hecho=str(row.get('id_hecho', '')),
                    nro_registro=str(row.get('nro_registro', '')),
                    ipp=str(row.get('ipp', '')),
                    relato=texto_relato,
                    jurisdiccion=clasificacion['jurisdiccion'],
                    calificacion=clasificacion['calificacion'],
                    modalidad=clasificacion['modalidad'],
                    victimas=clasificacion['victimas'],
                    lesionado=clasificacion['lesionado'],
                    imputados=clasificacion['imputados'],
                    edad=clasificacion['edad'],
                    armas=clasificacion['armas'],
                    lugar=clasificacion['lugar'],
                    tentativa=clasificacion['tentativa'],
                    observaciones=clasificacion['observaciones'],
                    metodo_clasificacion=metodo,
                    confianza_clasificacion=confianza,
                    tiempo_procesamiento=tiempo_ms
                )
                
                # Agregar otros campos del Excel si existen
                for col in df.columns:
                    if hasattr(hecho, col.lower()) and col.lower() not in ['id', 'planilla_id']:
                        try:
                            setattr(hecho, col.lower(), str(row[col]) if pd.notna(row[col]) else None)
                        except:
                            pass
                
                db.session.add(hecho)
                
                # Actualizar progreso
                planilla.registros_procesados = index + 1
                db.session.commit()
                
                # Log de progreso
                if (index + 1) % 10 == 0:
                    log = LogProcesamiento(
                        planilla_id=planilla_id,
                        nivel='INFO',
                        mensaje=f'Procesadas {index + 1} de {len(df)} filas',
                        metodo=metodo,
                        tiempo_ejecucion=tiempo_ms
                    )
                    db.session.add(log)
                    db.session.commit()
            
            # Marcar como completado
            planilla.estado = 'COMPLETADO'
            db.session.commit()
            
            # Limpiar proceso activo
            if planilla_id in procesos_activos:
                del procesos_activos[planilla_id]
                
    except Exception as e:
        logger.error(f"Error procesando archivo {planilla_id}: {str(e)}")
        try:
            planilla = Planilla.query.get(planilla_id)
            if planilla:
                planilla.estado = 'ERROR'
                planilla.observaciones = str(e)
                db.session.commit()
        except:
            pass

def calcular_tiempo_estimado(planilla):
    """Calcula tiempo estimado restante"""
    if planilla.estado != 'PROCESANDO':
        return 0
    
    restantes = planilla.total_registros - planilla.registros_procesados
    return restantes * 2  # 2 segundos por registro estimado

@clasificador_bp.route('/status/<int:planilla_id>')
def get_processing_status(planilla_id):
    """Obtiene el estado del procesamiento"""
    try:
        planilla = Planilla.query.get_or_404(planilla_id)
        
        return jsonify({
            'planilla_id': planilla_id,
            'estado': planilla.estado,
            'total_registros': planilla.total_registros,
            'registros_procesados': planilla.registros_procesados,
            'progreso': (planilla.registros_procesados / planilla.total_registros * 100) if planilla.total_registros > 0 else 0,
            'tiempo_estimado': calcular_tiempo_estimado(planilla),
            'nombre_archivo': planilla.nombre_archivo
        })
        
    except Exception as e:
        logger.error(f"Error obteniendo estado: {str(e)}")
        return jsonify({'error': f'Error: {str(e)}'}), 500

@clasificador_bp.route('/cancel/<int:planilla_id>', methods=['POST'])
def cancel_processing(planilla_id):
    """Cancelar procesamiento"""
    try:
        # Marcar como cancelado en la estructura de control
        if planilla_id in procesos_activos:
            procesos_activos[planilla_id]['cancelado'] = True
        
        # Actualizar estado en base de datos
        planilla = Planilla.query.get(planilla_id)
        if planilla:
            planilla.estado = 'CANCELADO'
            db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Procesamiento cancelado'
        })
        
    except Exception as e:
        logger.error(f"Error cancelando procesamiento: {str(e)}")
        return jsonify({'error': f'Error: {str(e)}'}), 500

@clasificador_bp.route('/generate-excel/<int:planilla_id>')
def generate_excel(planilla_id):
    """Generar archivo Excel con resultados clasificados"""
    try:
        planilla = Planilla.query.get_or_404(planilla_id)
        
        if planilla.estado != 'COMPLETADO':
            return jsonify({'error': 'La planilla no ha sido procesada completamente'}), 400
        
        # Obtener hechos delictivos clasificados
        hechos = HechoDelictivo.query.filter_by(planilla_id=planilla_id).all()
        
        if not hechos:
            return jsonify({'error': 'No hay datos clasificados para exportar'}), 404
        
        # Preparar datos para Excel
        data = []
        for hecho in hechos:
            data.append({
                'id_hecho': hecho.id_hecho,
                'nro_registro': hecho.nro_registro,
                'ipp': hecho.ipp,
                'fecha_carga': hecho.fecha_carga.isoformat() if hecho.fecha_carga else '',
                'fecha_hecho': hecho.fecha_hecho.isoformat() if hecho.fecha_hecho else '',
                'hora_hecho': hecho.hora_hecho.isoformat() if hecho.hora_hecho else '',
                'comisaria': hecho.comisaria,
                'dependencia': hecho.dependencia,
                'localidad': hecho.localidad,
                'barrio': hecho.barrio,
                'direccion': hecho.direccion,
                'latitud': float(hecho.latitud) if hecho.latitud else '',
                'longitud': float(hecho.longitud) if hecho.longitud else '',
                'relato': hecho.relato,
                'jurisdiccion': hecho.jurisdiccion,
                'calificacion': hecho.calificacion,
                'modalidad': hecho.modalidad,
                'victimas': hecho.victimas,
                'lesionado': hecho.lesionado,
                'imputados': hecho.imputados,
                'edad': hecho.edad,
                'armas': hecho.armas,
                'lugar': hecho.lugar,
                'tentativa': hecho.tentativa,
                'observaciones': hecho.observaciones,
                'metodo_clasificacion': hecho.metodo_clasificacion,
                'confianza_clasificacion': float(hecho.confianza_clasificacion) if hecho.confianza_clasificacion else '',
                'tiempo_procesamiento_ms': hecho.tiempo_procesamiento
            })
        
        # Crear DataFrame
        df = pd.DataFrame(data)
        
        # Crear workbook con formato
        wb = Workbook()
        ws = wb.active
        ws.title = "Delitos Clasificados"
        
        # Escribir headers
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Escribir datos
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en buffer
        import io
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"clasificado_{planilla.nombre_archivo}"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error generando Excel: {str(e)}")
        return jsonify({'error': f'Error generando Excel: {str(e)}'}), 500

@clasificador_bp.route('/planillas')
def get_planillas():
    """Obtener lista de planillas"""
    try:
        planillas = Planilla.query.order_by(Planilla.fecha_subida.desc()).all()
        
        result = []
        for planilla in planillas:
            progreso = 0
            if planilla.total_registros > 0:
                progreso = (planilla.registros_procesados / planilla.total_registros) * 100
            
            result.append({
                'id': planilla.id,
                'nombre_archivo': planilla.nombre_archivo,
                'fecha_subida': planilla.fecha_subida.isoformat() if planilla.fecha_subida else None,
                'total_registros': planilla.total_registros,
                'registros_procesados': planilla.registros_procesados or 0,
                'estado': planilla.estado,
                'progreso': round(progreso, 1),
                'partido': planilla.partido,
                'observaciones': planilla.observaciones
            })
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Error obteniendo planillas: {str(e)}")
        return jsonify({'error': f'Error obteniendo planillas: {str(e)}'}), 500

@clasificador_bp.route('/test')
def test_clasificador():
    """Endpoint de prueba para verificar el funcionamiento"""
    try:
        # Probar clasificación con texto de ejemplo
        texto_prueba = "El día 15 de mayo de 2024, en horas de la madrugada, se produjo un robo a mano armada en la vía pública donde un sujeto de sexo masculino interceptó a la victima con un arma de fuego y le sustrajo sus pertenencias."
        
        clasificacion, metodo, tiempo_ms, confianza = clasificar_fila_con_cascada(texto_prueba, 1)
        
        return jsonify({
            'success': True,
            'texto_prueba': texto_prueba,
            'clasificacion': clasificacion,
            'metodo': metodo,
            'tiempo_ms': tiempo_ms,
            'confianza': confianza,
            'message': 'Sistema de clasificación funcionando correctamente'
        })
        
    except Exception as e:
        logger.error(f"Error en test: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Error en el sistema de clasificación'
        }), 500

