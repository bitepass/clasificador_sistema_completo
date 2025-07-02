from flask import Blueprint, request, jsonify, current_app, send_file
from werkzeug.utils import secure_filename
import pandas as pd
import os
import json
import time
import threading
from datetime import datetime, date
import traceback
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import google.generativeai as genai
import openai
import re

from src.models.models import db, Planilla, HechoDelictivo, LogProcesamiento

clasificador_bp = Blueprint('clasificador', __name__)

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Variables globales para control de proceso
procesos_activos = {}

# Configurar APIs
def configurar_apis():
    """Configura las APIs de Gemini y OpenAI"""
    try:
        gemini_key = os.getenv('GEMINI_API_KEY')
        if gemini_key:
            genai.configure(api_key=gemini_key)
        
        openai_key = os.getenv('OPENAI_API_KEY')
        if openai_key:
            openai.api_key = openai_key
    except Exception as e:
        logger.error(f"Error configurando APIs: {str(e)}")

# Categorías de clasificación
CATEGORIAS_CLASIFICACION = {
    'JURISDICCION': ['URBANA', 'RURAL', 'MIXTA', 'OTROS'],
    'CALIFICACION': [
        'HOMICIDIO_SIMPLE', 'FEMICIDIO', 'INTRAFAMILIAR', 'EN_RINA', 'EN_OCASION_DE_ROBO', 'AJUSTE_DE_CUENTAS',
        'ROBO_ESCRUCHE', 'ROBO_ENTRADERA', 'ROBO_ASALTO_FINCA', 'ROBO_ASALTO_VIA_PUBLICA', 'ROBO_ASALTO_COMERCIO',
        'ROBO_MOTOCHORROS', 'ROBO_ROBACABLES', 'ROBO_SIMPLE', 'ROBO_ROBARRUEDAS', 'ROBO_ROMPEVIDRIOS',
        'ROBO_ARREBATADOR', 'ROBO_BICICLETA', 'ROBO_CHOFERES_REPARTIDORES',
        'ENFRENTAMIENTOS_EN_OCASION_DE_ROBO', 'ENFRENTAMIENTOS_AJUSTE_DE_CUENTAS', 'ENFRENTAMIENTOS_PROCEDIMIENTO_POLICIAL',
        'ENFRENTAMIENTOS_EN_RINA', 'ENFRENTAMIENTOS_BANDAS_ANTAGONICAS',
        'SUSTRACCION_AUTOMOTOR_LEVANTAMIENTO', 'SUSTRACCION_AUTOMOTOR_ASALTO', 'SUSTRACCION_MOTOVEHICULO_LEVANTAMIENTO',
        'SUSTRACCION_MOTOVEHICULO_ASALTO', 'LESIONES_ARMA_FUEGO', 'LESIONES_ARMA_BLANCA', 'LESIONES_ARMA_IMPROPIA',
        'USURPACION', 'ABUSO_SEXUAL_SIMPLE', 'ABUSO_SEXUAL_ACCESO_CARNAL', 'LEY_23737_TENENCIA', 'LEY_23737_CONSUMO',
        'LEY_23737_COMERCIALIZACION', 'LEY_23737_SIEMBRA', 'ABIGEATO', 'ESTAFAS_MARKETPLACE', 'ESTAFAS_WHATSAPP',
        'ESTAFAS_CUENTO_DEL_TIO', 'ESTAFAS_OTROS', 'ABUSO_DE_ARMAS', 'TENENCIA_DE_ARMAS', 'PORTACION_DE_ARMAS',
        'ENCUBRIMIENTO_VIA_PUBLICA', 'ENCUBRIMIENTO_TALLER', 'ENCUBRIMIENTO_DOMICILIO_PARTICULAR', 'NINGUNO_DE_INTERES'
    ],
    'MODALIDAD': [
        'MODALIDAD_HOMICIDIO_SIMPLE', 'MODALIDAD_FEMICIDIO', 'MODALIDAD_INTRAFAMILIAR', 'MODALIDAD_EN_RINA',
        'MODALIDAD_EN_OCASION_DE_ROBO', 'MODALIDAD_AJUSTE_DE_CUENTAS', 'MODALIDAD_ROBO_ESCRUCHE',
        'MODALIDAD_ROBO_ENTRADERA', 'MODALIDAD_ROBO_ASALTO_FINCA', 'MODALIDAD_ROBO_ASALTO_VIA_PUBLICA',
        'MODALIDAD_ROBO_ASALTO_COMERCIO', 'MODALIDAD_ROBO_MOTOCHORROS', 'MODALIDAD_ROBO_ROBACABLES',
        'MODALIDAD_ROBO_SIMPLE', 'MODALIDAD_ROBO_ROBARRUEDAS', 'MODALIDAD_ROBO_ROMPEVIDRIOS',
        'MODALIDAD_ROBO_ARREBATADOR', 'MODALIDAD_ROBO_BICICLETA', 'MODALIDAD_ROBO_CHOFERES_REPARTIDORES',
        'MODALIDAD_ENFRENTAMIENTOS_EN_OCASION_DE_ROBO', 'MODALIDAD_ENFRENTAMIENTOS_AJUSTE_DE_CUENTAS',
        'MODALIDAD_ENFRENTAMIENTOS_PROCEDIMIENTO_POLICIAL', 'MODALIDAD_ENFRENTAMIENTOS_EN_RINA',
        'MODALIDAD_ENFRENTAMIENTOS_BANDAS_ANTAGONICAS', 'MODALIDAD_SUSTRACCION_AUTOMOTOR_LEVANTAMIENTO',
        'MODALIDAD_SUSTRACCION_AUTOMOTOR_ASALTO', 'MODALIDAD_SUSTRACCION_MOTOVEHICULO_LEVANTAMIENTO',
        'MODALIDAD_SUSTRACCION_MOTOVEHICULO_ASALTO', 'MODALIDAD_LESIONES_ARMA_FUEGO',
        'MODALIDAD_LESIONES_ARMA_BLANCA', 'MODALIDAD_LESIONES_ARMA_IMPROPIA', 'OTROS'
    ],
    'VICTIMAS': ['FEMENINO', 'MASCULINO', 'AMBOS', 'OTROS'],
    'LESIONADO': ['SI', 'NO', 'OTROS'],
    'IMPUTADOS': ['FEMENINO', 'MASCULINO', 'AMBOS', 'OTROS'],
    'EDAD': ['MAYOR', 'MENOR', 'AMBOS', 'OTROS'],
    'ARMAS': ['FUEGO', 'BLANCA', 'IMPROPIA', 'NINGUNA', 'OTROS'],
    'LUGAR': ['FINCA', 'VIA_PUBLICA', 'COMERCIO', 'ESTABLECIMIENTO_EDUCATIVO', 'OTROS'],
    'TENTATIVA': ['SI', 'NO', 'OTROS']
}

def clasificar_texto_con_gemini(texto):
    """Clasificar texto usando Gemini API"""
    try:
        model = genai.GenerativeModel('gemini-pro')
        
        prompt = f"""
        Analiza el siguiente texto de un hecho delictivo y clasifícalo según estas categorías exactas:

        JURISDICCION: {', '.join(CATEGORIAS_CLASIFICACION['JURISDICCION'])}
        CALIFICACION: {', '.join(CATEGORIAS_CLASIFICACION['CALIFICACION'])}
        MODALIDAD: {', '.join(CATEGORIAS_CLASIFICACION['MODALIDAD'])}
        VICTIMAS: {', '.join(CATEGORIAS_CLASIFICACION['VICTIMAS'])}
        LESIONADO: {', '.join(CATEGORIAS_CLASIFICACION['LESIONADO'])}
        IMPUTADOS: {', '.join(CATEGORIAS_CLASIFICACION['IMPUTADOS'])}
        EDAD: {', '.join(CATEGORIAS_CLASIFICACION['EDAD'])}
        ARMAS: {', '.join(CATEGORIAS_CLASIFICACION['ARMAS'])}
        LUGAR: {', '.join(CATEGORIAS_CLASIFICACION['LUGAR'])}
        TENTATIVA: {', '.join(CATEGORIAS_CLASIFICACION['TENTATIVA'])}

        Texto a analizar: {texto}

        Responde SOLO en formato JSON válido con estas claves exactas:
        {{
            "jurisdiccion": "valor",
            "calificacion": "valor",
            "modalidad": "valor",
            "victimas": "valor",
            "lesionado": "valor",
            "imputados": "valor",
            "edad": "valor",
            "armas": "valor",
            "lugar": "valor",
            "tentativa": "valor",
            "observaciones": "texto libre con detalles relevantes"
        }}
        """
        
        response = model.generate_content(prompt)
        resultado = json.loads(response.text.strip())
        return resultado, 'GEMINI'
        
    except Exception as e:
        logger.error(f"Error en clasificación con Gemini: {str(e)}")
        return None, 'ERROR_GEMINI'

def clasificar_con_openai_alternativa(texto):
    """Clasificar texto usando OpenAI como alternativa"""
    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Eres un experto en clasificación de delitos. Responde solo en formato JSON válido."},
                {"role": "user", "content": f"""
                Clasifica este hecho delictivo según las categorías especificadas:
                
                JURISDICCION: {', '.join(CATEGORIAS_CLASIFICACION['JURISDICCION'])}
                CALIFICACION: {', '.join(CATEGORIAS_CLASIFICACION['CALIFICACION'])}
                MODALIDAD: {', '.join(CATEGORIAS_CLASIFICACION['MODALIDAD'])}
                VICTIMAS: {', '.join(CATEGORIAS_CLASIFICACION['VICTIMAS'])}
                LESIONADO: {', '.join(CATEGORIAS_CLASIFICACION['LESIONADO'])}
                IMPUTADOS: {', '.join(CATEGORIAS_CLASIFICACION['IMPUTADOS'])}
                EDAD: {', '.join(CATEGORIAS_CLASIFICACION['EDAD'])}
                ARMAS: {', '.join(CATEGORIAS_CLASIFICACION['ARMAS'])}
                LUGAR: {', '.join(CATEGORIAS_CLASIFICACION['LUGAR'])}
                TENTATIVA: {', '.join(CATEGORIAS_CLASIFICACION['TENTATIVA'])}

                Texto: {texto}

                Formato de respuesta JSON:
                {{
                    "jurisdiccion": "valor",
                    "calificacion": "valor",
                    "modalidad": "valor",
                    "victimas": "valor",
                    "lesionado": "valor",
                    "imputados": "valor",
                    "edad": "valor",
                    "armas": "valor",
                    "lugar": "valor",
                    "tentativa": "valor",
                    "observaciones": "detalles relevantes"
                }}
                """}
            ],
            temperature=0.1
        )
        
        resultado = json.loads(response.choices[0].message.content.strip())
        return resultado, 'OPENAI'
        
    except Exception as e:
        logger.error(f"Error en clasificación con OpenAI: {str(e)}")
        return None, 'ERROR_OPENAI'

def clasificacion_basada_en_reglas(texto):
    """Clasificación básica basada en reglas como último recurso"""
    try:
        texto_lower = texto.lower()
        
        # Reglas básicas de clasificación
        clasificacion = {
            'jurisdiccion': 'OTROS',
            'calificacion': 'OTROS',
            'modalidad': 'OTROS',
            'victimas': 'OTROS',
            'lesionado': 'OTROS',
            'imputados': 'OTROS',
            'edad': 'OTROS',
            'armas': 'NINGUNA',
            'lugar': 'OTROS',
            'tentativa': 'NO',
            'observaciones': 'Clasificación automática por reglas'
        }
        
        # Reglas para calificación
        if any(word in texto_lower for word in ['homicidio', 'asesinato', 'muerte']):
            clasificacion['calificacion'] = 'HOMICIDIO_SIMPLE'
        elif any(word in texto_lower for word in ['robo', 'asalto']):
            clasificacion['calificacion'] = 'ROBO_SIMPLE'
        elif any(word in texto_lower for word in ['lesiones', 'herido']):
            clasificacion['calificacion'] = 'LESIONES_ARMA_IMPROPIA'
        
        # Reglas para armas
        if any(word in texto_lower for word in ['arma de fuego', 'pistola', 'revolver', 'bala']):
            clasificacion['armas'] = 'FUEGO'
        elif any(word in texto_lower for word in ['cuchillo', 'navaja', 'arma blanca']):
            clasificacion['armas'] = 'BLANCA'
        
        # Reglas para lugar
        if any(word in texto_lower for word in ['calle', 'vía pública', 'vereda']):
            clasificacion['lugar'] = 'VIA_PUBLICA'
        elif any(word in texto_lower for word in ['casa', 'domicilio', 'vivienda']):
            clasificacion['lugar'] = 'FINCA'
        elif any(word in texto_lower for word in ['comercio', 'negocio', 'tienda']):
            clasificacion['lugar'] = 'COMERCIO'
        
        return clasificacion, 'REGLAS'
        
    except Exception as e:
        logger.error(f"Error en clasificación por reglas: {str(e)}")
        return None, 'ERROR_REGLAS'

def clasificacion_por_defecto():
    """Clasificación por defecto cuando todo falla"""
    return {
        'jurisdiccion': 'OTROS',
        'calificacion': 'NINGUNO_DE_INTERES',
        'modalidad': 'OTROS',
        'victimas': 'OTROS',
        'lesionado': 'OTROS',
        'imputados': 'OTROS',
        'edad': 'OTROS',
        'armas': 'NINGUNA',
        'lugar': 'OTROS',
        'tentativa': 'NO',
        'observaciones': 'Sin información suficiente para clasificar'
    }, 'DEFECTO'

def limpiar_clasificacion(clasificacion):
    """Valida y limpia la clasificación obtenida"""
    if not isinstance(clasificacion, dict):
        return None
    
    clasificacion_limpia = {}
    
    for categoria, valor in clasificacion.items():
        if categoria in CATEGORIAS_CLASIFICACION:
            if valor in CATEGORIAS_CLASIFICACION[categoria]:
                clasificacion_limpia[categoria] = valor
            else:
                clasificacion_limpia[categoria] = 'OTROS'
        elif categoria == 'observaciones':
            clasificacion_limpia[categoria] = str(valor)[:500]  # Limitar longitud
    
    # Asegurar que todas las categorías estén presentes
    for categoria in CATEGORIAS_CLASIFICACION.keys():
        if categoria not in clasificacion_limpia:
            clasificacion_limpia[categoria] = 'OTROS'
    
    if 'observaciones' not in clasificacion_limpia:
        clasificacion_limpia['observaciones'] = ''
    
    return clasificacion_limpia

def clasificar_fila_con_cascada(texto, fila_num):
    """Clasifica una fila usando el sistema de cascada"""
    inicio_tiempo = time.time()
    
    # Si no hay texto, usar clasificación por defecto
    if not texto or str(texto).strip() == '' or pd.isna(texto):
        clasificacion, metodo = clasificacion_por_defecto()
        tiempo_ms = int((time.time() - inicio_tiempo) * 1000)
        return clasificacion, metodo, tiempo_ms, 0.0
    
    # Intentar con Gemini
    clasificacion, metodo = clasificar_texto_con_gemini(str(texto))
    if clasificacion:
        clasificacion_limpia = limpiar_clasificacion(clasificacion)
        if clasificacion_limpia:
            tiempo_ms = int((time.time() - inicio_tiempo) * 1000)
            return clasificacion_limpia, metodo, tiempo_ms, 0.9
    
    # Intentar con OpenAI
    clasificacion, metodo = clasificar_con_openai_alternativa(str(texto))
    if clasificacion:
        clasificacion_limpia = limpiar_clasificacion(clasificacion)
        if clasificacion_limpia:
            tiempo_ms = int((time.time() - inicio_tiempo) * 1000)
            return clasificacion_limpia, metodo, tiempo_ms, 0.8
    
    # Intentar con reglas
    clasificacion, metodo = clasificacion_basada_en_reglas(str(texto))
    if clasificacion:
        clasificacion_limpia = limpiar_clasificacion(clasificacion)
        if clasificacion_limpia:
            tiempo_ms = int((time.time() - inicio_tiempo) * 1000)
            return clasificacion_limpia, metodo, tiempo_ms, 0.5
    
    # Usar clasificación por defecto
    clasificacion, metodo = clasificacion_por_defecto()
    tiempo_ms = int((time.time() - inicio_tiempo) * 1000)
    return clasificacion, metodo, tiempo_ms, 0.1

@clasificador_bp.route('/upload', methods=['POST'])
def upload_file():
    """Endpoint para subir archivo Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No se encontró archivo'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Solo se permiten archivos Excel (.xlsx, .xls)'}), 400
        
        # Guardar archivo
        filename = secure_filename(file.filename)
        upload_path = os.path.join('/tmp', filename)
        file.save(upload_path)
        
        # Leer Excel para obtener información básica
        df = pd.read_excel(upload_path)
        total_filas = len(df)
        
        # Crear registro en base de datos
        planilla = Planilla(
            nombre_archivo=filename,
            total_registros=total_filas,
            estado='SUBIDO'
        )
        db.session.add(planilla)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'planilla_id': planilla.id,
            'filename': filename,
            'total_filas': total_filas,
            'message': 'Archivo subido correctamente'
        })
        
    except Exception as e:
        logger.error(f"Error subiendo archivo: {str(e)}")
        return jsonify({'error': f'Error procesando archivo: {str(e)}'}), 500

@clasificador_bp.route('/process/<int:planilla_id>', methods=['POST'])
def process_file(planilla_id):
    """Endpoint para procesar archivo con clasificación"""
    try:
        configurar_apis()
        
        planilla = Planilla.query.get_or_404(planilla_id)
        
        if planilla.estado == 'PROCESANDO':
            return jsonify({'error': 'El archivo ya se está procesando'}), 400
        
        # Marcar como procesando
        planilla.estado = 'PROCESANDO'
        planilla.registros_procesados = 0
        db.session.commit()
        
        # Iniciar procesamiento en hilo separado
        thread = threading.Thread(
            target=procesar_archivo_async,
            args=(planilla_id,)
        )
        thread.daemon = True
        thread.start()
        
        # Guardar referencia del proceso
        procesos_activos[planilla_id] = {
            'thread': thread,
            'cancelado': False,
            'inicio': datetime.now()
        }
        
        return jsonify({
            'success': True,
            'message': 'Procesamiento iniciado',
            'planilla_id': planilla_id
        })
        
    except Exception as e:
        logger.error(f"Error iniciando procesamiento: {str(e)}")
        return jsonify({'error': f'Error: {str(e)}'}), 500

def procesar_archivo_async(planilla_id):
    """Procesa el archivo de forma asíncrona"""
    try:
        with current_app.app_context():
            planilla = Planilla.query.get(planilla_id)
            if not planilla:
                return
            
            # Leer archivo Excel
            upload_path = os.path.join('/tmp', planilla.nombre_archivo)
            df = pd.read_excel(upload_path)
            
            # Procesar cada fila
            for index, row in df.iterrows():
                # Verificar si fue cancelado
                if planilla_id in procesos_activos and procesos_activos[planilla_id]['cancelado']:
                    planilla.estado = 'CANCELADO'
                    db.session.commit()
                    return
                
                # Obtener texto para clasificar (columna 'relato' o similar)
                texto_relato = ''
                for col in ['relato', 'descripcion', 'detalle', 'observaciones']:
                    if col in df.columns:
                        texto_relato = str(row[col]) if pd.notna(row[col]) else ''
                        break
                
                # Clasificar fila
                clasificacion, metodo, tiempo_ms, confianza = clasificar_fila_con_cascada(texto_relato, index)
                
                # Crear registro en base de datos
                hecho = HechoDelictivo(
                    planilla_id=planilla_id,
                    id_hecho=str(row.get('id_hecho', '')),
                    nro_registro=str(row.get('nro_registro', '')),
                    ipp=str(row.get('ipp', '')),
                    relato=texto_relato,
                    jurisdiccion=clasificacion['jurisdiccion'],
                    calificacion=clasificacion['calificacion'],
                    modalidad=clasificacion['modalidad'],
                    victimas=clasificacion['victimas'],
                    lesionado=clasificacion['lesionado'],
                    imputados=clasificacion['imputados'],
                    edad=clasificacion['edad'],
                    armas=clasificacion['armas'],
                    lugar=clasificacion['lugar'],
                    tentativa=clasificacion['tentativa'],
                    observaciones=clasificacion['observaciones'],
                    metodo_clasificacion=metodo,
                    confianza_clasificacion=confianza,
                    tiempo_procesamiento=tiempo_ms
                )
                
                # Agregar otros campos del Excel si existen
                for col in df.columns:
                    if hasattr(hecho, col.lower()) and col.lower() not in ['id', 'planilla_id']:
                        try:
                            setattr(hecho, col.lower(), str(row[col]) if pd.notna(row[col]) else None)
                        except:
                            pass
                
                db.session.add(hecho)
                
                # Actualizar progreso
                planilla.registros_procesados = index + 1
                db.session.commit()
                
                # Log de progreso
                if (index + 1) % 10 == 0:
                    log = LogProcesamiento(
                        planilla_id=planilla_id,
                        nivel='INFO',
                        mensaje=f'Procesadas {index + 1} de {len(df)} filas',
                        metodo=metodo,
                        tiempo_ejecucion=tiempo_ms
                    )
                    db.session.add(log)
                    db.session.commit()
            
            # Marcar como completado
            planilla.estado = 'COMPLETADO'
            db.session.commit()
            
            # Limpiar proceso activo
            if planilla_id in procesos_activos:
                del procesos_activos[planilla_id]
                
    except Exception as e:
        logger.error(f"Error procesando archivo {planilla_id}: {str(e)}")
        try:
            planilla = Planilla.query.get(planilla_id)
            if planilla:
                planilla.estado = 'ERROR'
                planilla.observaciones = str(e)
                db.session.commit()
        except:
            pass

@clasificador_bp.route('/status/<int:planilla_id>')
def get_status(planilla_id):
    """Obtiene el estado del procesamiento"""
    try:
        planilla = Planilla.query.get_or_404(planilla_id)
        
        return jsonify({
            'planilla_id': planilla_id,
            'estado': planilla.estado,
            'total_registros': planilla.total_registros,
            'registros_procesados': planilla.registros_procesados,
            'progreso': (planilla.registros_procesados / planilla.total_registros * 100) if planilla.total_registros > 0 else 0,
            'tiempo_estimado': calcular_tiempo_estimado(planilla),
            'puede_cancelar': planilla_id in procesos_activos
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def calcular_tiempo_estimado(planilla):
    """Calcula tiempo estimado restante"""
    if planilla.estado != 'PROCESANDO' or planilla.registros_procesados == 0:
        return 0
    
    if planilla.id in procesos_activos:
        tiempo_transcurrido = (datetime.now() - procesos_activos[planilla.id]['inicio']).total_seconds()
        tiempo_por_registro = tiempo_transcurrido / planilla.registros_procesados
        registros_restantes = planilla.total_registros - planilla.registros_procesados
        return int(registros_restantes * tiempo_por_registro)
    
    return 0

@clasificador_bp.route('/cancel/<int:planilla_id>', methods=['POST'])
def cancel_process(planilla_id):
    """Cancela el procesamiento"""
    try:
        if planilla_id in procesos_activos:
            procesos_activos[planilla_id]['cancelado'] = True
            return jsonify({'success': True, 'message': 'Proceso cancelado'})
        else:
            return jsonify({'error': 'No hay proceso activo para cancelar'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@clasificador_bp.route('/generate-excel/<int:planilla_id>')
def generate_excel(planilla_id):
    """Genera archivo Excel con resultados"""
    try:
        planilla = Planilla.query.get_or_404(planilla_id)
        
        if planilla.estado != 'COMPLETADO':
            return jsonify({'error': 'El procesamiento no ha terminado'}), 400
        
        # Obtener datos procesados
        hechos = HechoDelictivo.query.filter_by(planilla_id=planilla_id).all()
        
        if not hechos:
            return jsonify({'error': 'No hay datos procesados'}), 400
        
        # Crear Excel con formato
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados Clasificados"
        
        # Color púrpura para columnas clasificadas
        purple_fill = PatternFill(start_color="C0B8E8", end_color="C0B8E8", fill_type="solid")
        
        # Encabezados
        headers = [
            'id_hecho', 'nro_registro', 'ipp', 'fecha_carga', 'fecha_hecho', 'hora_hecho',
            'comisaria', 'dependencia', 'localidad', 'barrio', 'direccion', 'relato',
            'JURISDICCION', 'CALIFICACION', 'MODALIDAD', 'VICTIMAS', 'LESIONADO',
            'IMPUTADOS', 'EDAD', 'ARMAS', 'LUGAR', 'TENTATIVA', 'OBSERVACIONES'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            # Aplicar color púrpura a columnas clasificadas
            if header in ['JURISDICCION', 'CALIFICACION', 'MODALIDAD', 'VICTIMAS', 'LESIONADO',
                         'IMPUTADOS', 'EDAD', 'ARMAS', 'LUGAR', 'TENTATIVA', 'OBSERVACIONES']:
                cell.fill = purple_fill
        
        # Escribir datos
        for row_num, hecho in enumerate(hechos, 2):
            data = [
                hecho.id_hecho, hecho.nro_registro, hecho.ipp, hecho.fecha_carga, hecho.fecha_hecho,
                hecho.hora_hecho, hecho.comisaria, hecho.dependencia, hecho.localidad, hecho.barrio,
                hecho.direccion, hecho.relato, hecho.jurisdiccion, hecho.calificacion, hecho.modalidad,
                hecho.victimas, hecho.lesionado, hecho.imputados, hecho.edad, hecho.armas,
                hecho.lugar, hecho.tentativa, hecho.observaciones
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                # Aplicar color púrpura a columnas clasificadas
                if col > 12:  # Columnas clasificadas empiezan en la 13
                    cell.fill = purple_fill
        
        # Guardar archivo
        output_filename = f"clasificado_{planilla.nombre_archivo}"
        output_path = os.path.join('/tmp', output_filename)
        wb.save(output_path)
        
        # Actualizar planilla con archivo resultado
        planilla.archivo_resultado = output_filename
        db.session.commit()
        
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error generando Excel: {str(e)}")
        return jsonify({'error': f'Error generando archivo: {str(e)}'}), 500

@clasificador_bp.route('/planillas')
def list_planillas():
    """Lista todas las planillas"""
    try:
        planillas = Planilla.query.order_by(Planilla.fecha_subida.desc()).all()
        return jsonify([planilla.to_dict() for planilla in planillas])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@clasificador_bp.route('/test')
def test_clasificador():
    """Test del sistema de clasificación"""
    try:
        configurar_apis()
        
        texto_prueba = "Se produjo un robo a mano armada en la vía pública donde una persona de sexo masculino sustrajo el teléfono celular a una mujer utilizando un arma de fuego"
        
        clasificacion, metodo, tiempo_ms, confianza = clasificar_fila_con_cascada(texto_prueba, 1)
        
        return jsonify({
            'success': True,
            'texto_prueba': texto_prueba,
            'clasificacion': clasificacion,
            'metodo': metodo,
            'tiempo_ms': tiempo_ms,
            'confianza': confianza
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

